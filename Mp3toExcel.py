import mutagen,xlrd, glob,re,openpyxl,os,pygal
#from mutagen.easyid3 import EasyID3
from os import walk
from pprint import pprint 
from tinytag import TinyTag, TinyTagException
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from mp3_tagger import MP3File
from string import ascii_uppercase
from mutagen.mp3 import MP3
list = os.listdir('C:\\Users\\mrdrj\\Desktop\\sdf\\') # directory path of files
number_files = len(list) +1
from openpyxl.workbook import Workbook

tracks= []
gettags =[]
getit = []

def ExtractMP3TagtoExcel():
 
    for root, dirs, files, in os.walk ('C:\\Users\\mrdrj\\Desktop\\sdf\\'):
        for name in files:
            if name.endswith(('.mp3','.m4a','.flac','.alac')):

                tracks.append(name) #Add Media Files

                try:

                    temp_track = TinyTag.get(root + '\\' + name)
                    mp3 = MP3File(root + '\\' + name)
                    #tags = mp3.get_tags()
                    #print(root, '-',temp_track.artist, '-', temp_track.title)

                    gettags2 = [temp_track.album, temp_track.albumartist, temp_track.artist, temp_track.audio_offset,
                                temp_track.bitrate, temp_track.comment, temp_track.composer, temp_track.disc,
                                temp_track.disc_total, temp_track.duration, temp_track.filesize, temp_track.genre,
                                temp_track.samplerate, temp_track.title, temp_track.track, temp_track.track_total,
                                temp_track.year] #Add Tags to list
                   
               
    
                    for x in range(len(gettags2)):
                    #append slice of gettags2, containing the entire gettags2
                        gettags.append(gettags2[:])
                        #print(gettags2[x]) 
                except TinyTagException:
                    print('Error')

                
                os.chdir('C:\\Users\\mrdrj\\Desktop\\cqq\\CD 3\\')
                header = [u'album',u'albumartist' u'artist', u'audio_offset',u'bitrate', u'comment', u'composer', u'disc',u'disc_total',
                              u'duration', u'filesize', u'genre',u'samplerate', u'title', u'track', u'track_total',u'year']                             
                header2 = {u"album",u"albumartist" u"artist", u"audio_offset",u"bitrate", u"comment", u"composer", u"disc",u"disc_total",
                              u"duration", u"filesize", u"genre",u"samplerate", u"title", u"track", u"track_total",u"year"}

                
                new_date = gettags
                wb = Workbook()
                new_data = gettags
                dest_filename = '11empty_book11.xlsx'
                ws1 = wb.active
                ws1.title = "MP3 Tags"
                ws2 = wb.create_sheet(title="Set")
                ws1.append(header[:])
                
                tags = []
                
                   
                for row in new_data: # Number of Rows
                    tags.append(new_data[:]) #Add to Tag List
                    
                    ws1.append(row)
        
                        
                wb.save(filename=dest_filename)

                #wb.save(filename=dest_filename)


                    
                   #for row in range(1,new_data,16):
                   #for row in new_data: # Number of Rows
                    #tags.append(new_data[:]) #Add to Tag List
                                          
                        #for a in tags:
                           #if a not in tags:
                                #res.append(a) # Remove Duplicate Entries
                    #for row in range(len(new_data)6)
                                #ws1.append(str(a)) # Display in Rows in Excel
                                #wb.save(filename = dest_filename)
                                
                if len(row) == 0:
                    print('Completed ' + len(row))
             
            print(row)

ExtractMP3TagtoExcel()
